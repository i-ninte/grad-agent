"""Letter of recommendation tracker. xlsx-backed."""
from __future__ import annotations
import time
from pathlib import Path
from openpyxl import Workbook, load_workbook

PATH = Path(__file__).parent / "lor_log.xlsx"

HEADERS = ["date_added", "recommender", "email", "school",
           "program_id", "asked_at", "confirmed", "submitted_at", "notes"]


def _ensure():
    if PATH.exists(): return
    wb = Workbook(); ws = wb.active; ws.title = "lor"
    ws.append(HEADERS); wb.save(PATH)


def add(recommender: str, email: str = "", school: str = "",
        program_id: str = "", notes: str = "") -> int:
    _ensure()
    wb = load_workbook(PATH); ws = wb["lor"]
    ws.append([time.strftime("%Y-%m-%d"), recommender, email, school,
               program_id, "", "", "", notes])
    n = ws.max_row; wb.save(PATH); return n


def mark(recommender: str, school: str, field: str, value: str = "") -> bool:
    """field in {asked_at, confirmed, submitted_at, notes}"""
    _ensure()
    wb = load_workbook(PATH); ws = wb["lor"]
    idx = HEADERS.index(field)
    updated = False
    for row in ws.iter_rows(min_row=2):
        if (row[1].value or "").lower() == recommender.lower() and (row[3].value or "").lower() == school.lower():
            row[idx].value = value or time.strftime("%Y-%m-%d")
            updated = True
    if updated: wb.save(PATH)
    return updated


def outstanding() -> list[dict]:
    _ensure()
    wb = load_workbook(PATH, read_only=True); ws = wb["lor"]
    out = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[1]: continue
        row = dict(zip(HEADERS, r))
        if not row.get("submitted_at"):
            out.append(row)
    return out


def path() -> str:
    _ensure(); return str(PATH)
