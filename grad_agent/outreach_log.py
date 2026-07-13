"""Excel-backed outreach log. Every prof the agent surfaces or drafts to is
recorded here so tomorrow's run does not duplicate today's.

Dedup is keyed on the Semantic Scholar authorId (canonical and stable), with
a normalised-name fallback for rows written before the authorId column
existed or when S2 has no record.
"""
from __future__ import annotations
import re
import time
from openpyxl import Workbook, load_workbook

from . import config as _cfg


def _path():
    return _cfg.data_dir() / "outreach_log.xlsx"


HEADERS = [
    "date_added", "prof_name", "prof_lastname", "prof_email",
    "university", "program_target", "area", "paper_title",
    "paper_url", "project_matched", "match_score",
    "email_subject", "review_mail_sent_at", "sent_to_prof_at",
    "status", "notes",
    "s2_hindex", "s2_papers", "affiliation", "homepage",
    "recruiting_signal", "hook_attempts", "hook_verdict",
    "responded", "response_at", "outcome",
    "s2_author_id", "region",
    "fit_score", "fit_reason", "followup_sent_at",
]


SKIP_HEADERS = [
    "date", "source", "area", "prof_name", "stage", "reason",
    "paper_title", "arxiv_id", "resolution",
]


def _ensure():
    if not _path().exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "outreach"
        ws.append(HEADERS)
        ws2 = wb.create_sheet("skipped")
        ws2.append(SKIP_HEADERS)
        wb.save(_path())
        return
    # Older files predate the skipped sheet: add it in place.
    wb = load_workbook(_path())
    if "skipped" not in wb.sheetnames:
        ws2 = wb.create_sheet("skipped")
        ws2.append(SKIP_HEADERS)
        wb.save(_path())


def _norm_name(name: str) -> str:
    """Collapse a name to 'first-token last-token' so middle names and
    initials don't create duplicate keys (David Ifeoluwa Adelani == David Adelani)."""
    parts = re.findall(r"[a-z]+", (name or "").lower())
    if len(parts) >= 2:
        return f"{parts[0]} {parts[-1]}"
    return " ".join(parts)


def _read_rows() -> list[dict]:
    _ensure()
    wb = load_workbook(_path(), read_only=True)
    ws = wb["outreach"]
    header = [c.value for c in ws[1]]
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[1]:
            continue
        rows.append(dict(zip(header, r)))
    return rows


def already_contacted(prof_name: str, university: str = "",
                      author_id: str | None = None) -> bool:
    """Duplicate detection, in strength order:

    1. Both sides have an S2 authorId: equal -> duplicate, different ->
       PROVEN distinct people, even if the names collide. No name check.
    2. Either side lacks an authorId: fall back to normalised-name match
       (first + last token, so middle names and initials do not split keys).

    The university param is retained for API compatibility but is not part
    of the key (it was blank at log time too often to be useful)."""
    target_name = _norm_name(prof_name)
    for row in _read_rows():
        row_id = str(row.get("s2_author_id") or "").strip()
        if author_id and row_id:
            if row_id == str(author_id):
                return True
            continue  # both ids known and different: not the same person
        if target_name and _norm_name(str(row.get("prof_name") or "")) == target_name:
            return True
    return False


def add(**fields) -> int:
    """Append a row. Returns the new row number."""
    _ensure()
    fields.setdefault("date_added", time.strftime("%Y-%m-%d"))
    fields.setdefault("status", "drafted")
    wb = load_workbook(_path())
    ws = wb["outreach"]
    header = [c.value for c in ws[1]]
    # Extend the header in place if this file predates new columns.
    missing = [h for h in HEADERS if h not in header]
    for h in missing:
        ws.cell(row=1, column=len(header) + 1, value=h)
        header.append(h)
    ws.append([fields.get(h, "") for h in header])
    n = ws.max_row
    wb.save(_path())
    return n


def _set_fields(prof_name: str, values: dict) -> bool:
    """Set columns on every row matching the normalised prof name. Extends
    the header in place for columns the file predates."""
    _ensure()
    wb = load_workbook(_path())
    ws = wb["outreach"]
    header = [c.value for c in ws[1]]
    for col in values:
        if col not in header:
            ws.cell(row=1, column=len(header) + 1, value=col)
            header.append(col)
    target = _norm_name(prof_name)
    updated = False
    for row in ws.iter_rows(min_row=2):
        if _norm_name(str(row[1].value or "")) == target:
            for col, val in values.items():
                ws.cell(row=row[0].row, column=header.index(col) + 1, value=val)
            updated = True
    if updated:
        wb.save(_path())
    return updated


def mark_sent(prof_name: str, university: str = "") -> bool:
    return _set_fields(prof_name, {
        "sent_to_prof_at": time.strftime("%Y-%m-%d %H:%M"),
        "status": "sent",
    })


def mark_followup_sent(prof_name: str) -> bool:
    return _set_fields(prof_name, {
        "followup_sent_at": time.strftime("%Y-%m-%d"),
    })


def mark_response(prof_name: str, outcome: str, notes: str = "") -> bool:
    """outcome: positive | negative | no_response | interview | replied | other"""
    values = {
        "responded": "yes" if outcome != "no_response" else "no",
        "response_at": time.strftime("%Y-%m-%d"),
        "outcome": outcome,
    }
    if notes:
        values["notes"] = notes
    return _set_fields(prof_name, values)


def followups_due(days: int = 10) -> list[dict]:
    """Rows that were sent >= `days` ago, have no reply, and no follow-up yet."""
    import datetime as _dt
    cutoff = _dt.datetime.now() - _dt.timedelta(days=days)
    due = []
    for r in _read_rows():
        sent = str(r.get("sent_to_prof_at") or "").strip()
        if not sent:
            continue
        if str(r.get("responded") or "").lower() == "yes":
            continue
        if str(r.get("followup_sent_at") or "").strip():
            continue
        try:
            sent_dt = _dt.datetime.strptime(sent[:16], "%Y-%m-%d %H:%M")
        except ValueError:
            try:
                sent_dt = _dt.datetime.strptime(sent[:10], "%Y-%m-%d")
            except ValueError:
                continue
        if sent_dt <= cutoff:
            r["days_since_sent"] = (_dt.datetime.now() - sent_dt).days
            due.append(r)
    return due


def list_all(limit: int = 100) -> list[dict]:
    rows = _read_rows()
    return rows[-limit:]


def outcome_stats() -> dict:
    """Aggregate what the user tagged back into the log, so the agent can
    learn which areas and projects actually get responses."""
    rows = _read_rows()
    sent = [r for r in rows if r.get("sent_to_prof_at")]

    def _bucket(key: str) -> dict:
        agg: dict[str, dict] = {}
        for r in sent:
            k = str(r.get(key) or "unknown")
            a = agg.setdefault(k, {"sent": 0, "responded": 0, "positive": 0})
            a["sent"] += 1
            if str(r.get("responded") or "").lower() == "yes":
                a["responded"] += 1
            if str(r.get("outcome") or "").lower() in ("positive", "interview"):
                a["positive"] += 1
        for a in agg.values():
            a["response_rate"] = round(a["responded"] / a["sent"], 2) if a["sent"] else 0.0
        return agg

    return {
        "total_drafted": len(rows),
        "total_sent": len(sent),
        "total_responded": sum(1 for r in sent if str(r.get("responded") or "").lower() == "yes"),
        "by_area": _bucket("area"),
        "by_project": _bucket("project_matched"),
    }


def project_response_boost(min_sends: int = 3, boost: int = 4) -> dict[str, int]:
    """Score bonus per project name, derived from tagged outcomes. A project
    only earns a boost once it has at least `min_sends` sent emails, so one
    lucky reply doesn't distort matching."""
    stats = outcome_stats()["by_project"]
    out: dict[str, int] = {}
    for name, a in stats.items():
        if a["sent"] >= min_sends and a["response_rate"] > 0:
            out[name] = round(boost * a["response_rate"])
    return out


def add_skip(prof_name: str, stage: str, reason: str, source: str = "daily",
             area: str = "", paper_title: str = "", arxiv_id: str = "",
             resolution: str = "") -> int:
    """Persist a skipped lead so the mismatch is auditable after the fact.
    stage: identity | faculty-gate | region | dedup | no-papers"""
    _ensure()
    wb = load_workbook(_path())
    ws = wb["skipped"]
    ws.append([time.strftime("%Y-%m-%d"), source, area, prof_name, stage,
               reason, paper_title, arxiv_id, resolution])
    n = ws.max_row
    wb.save(_path())
    return n


def list_skipped(limit: int = 50) -> list[dict]:
    _ensure()
    wb = load_workbook(_path(), read_only=True)
    if "skipped" not in wb.sheetnames:
        return []
    ws = wb["skipped"]
    rows = [dict(zip(SKIP_HEADERS, r))
            for r in ws.iter_rows(min_row=2, values_only=True)
            if r and r[3]]
    return rows[-limit:]


def path() -> str:
    _ensure()
    return str(_path())
