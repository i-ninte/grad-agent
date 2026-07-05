"""Excel-backed outreach log. Every prof the agent surfaces or drafts to is
recorded here so tomorrow's run does not duplicate today's."""
from __future__ import annotations
import time
from pathlib import Path
from openpyxl import Workbook, load_workbook

PATH = Path(__file__).parent / "outreach_log.xlsx"

HEADERS = [
    "date_added", "prof_name", "prof_lastname", "prof_email",
    "university", "program_target", "area", "paper_title",
    "paper_url", "project_matched", "match_score",
    "email_subject", "review_mail_sent_at", "sent_to_prof_at",
    "status", "notes",
    # new v2 columns:
    "s2_hindex", "s2_papers", "affiliation", "homepage",
    "recruiting_signal", "hook_attempts", "hook_verdict",
    "responded", "response_at", "outcome",
]


def _ensure():
    if PATH.exists():
        return
    wb = Workbook()
    ws = wb.active
    ws.title = "outreach"
    ws.append(HEADERS)
    wb.save(PATH)


def _key(name: str, university: str) -> str:
    return f"{(name or '').strip().lower()}|{(university or '').strip().lower()}"


def already_contacted(prof_name: str, university: str) -> bool:
    _ensure()
    wb = load_workbook(PATH, read_only=True)
    ws = wb["outreach"]
    key = _key(prof_name, university)
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row: continue
        # columns: date, name, lastname, email, university, ...
        if _key(row[1] or "", row[4] or "") == key:
            return True
    return False


def add(**fields) -> int:
    """Append a row. Returns the new row number."""
    _ensure()
    fields.setdefault("date_added", time.strftime("%Y-%m-%d"))
    fields.setdefault("status", "drafted")
    wb = load_workbook(PATH)
    ws = wb["outreach"]
    row = [fields.get(h, "") for h in HEADERS]
    ws.append(row)
    n = ws.max_row
    wb.save(PATH)
    return n


def mark_sent(prof_name: str, university: str) -> bool:
    _ensure()
    wb = load_workbook(PATH)
    ws = wb["outreach"]
    key = _key(prof_name, university)
    updated = False
    for row in ws.iter_rows(min_row=2):
        if _key(row[1].value or "", row[4].value or "") == key:
            row[HEADERS.index("sent_to_prof_at")].value = time.strftime("%Y-%m-%d %H:%M")
            row[HEADERS.index("status")].value = "sent"
            updated = True
    if updated: wb.save(PATH)
    return updated


def list_all(limit: int = 100) -> list[dict]:
    _ensure()
    wb = load_workbook(PATH, read_only=True)
    ws = wb["outreach"]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    out = []
    for r in rows[-limit:]:
        out.append(dict(zip(HEADERS, r)))
    return out


def path() -> str:
    _ensure()
    return str(PATH)
